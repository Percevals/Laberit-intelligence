#!/usr/bin/env python3
"""
Simplified analysis of MasterDatabaseV3.1.xlsx without external dependencies.
This version uses openpyxl which should be available, or falls back to CSV if needed.
"""

import csv
import json
from collections import Counter, defaultdict
from datetime import datetime
import statistics

# DII 4.0 Model Definitions
DII_4_MODELS = {
    "A-RC": "Recovery Optimized for All Threats",
    "C-RY": "Recovery Optimized for Cyber Threats",
    "A-RS": "Resilience Optimized for All Threats", 
    "C-RS": "Resilience Optimized for Cyber Threats",
    "A-RG": "Response Optimized for All Threats",
    "C-RG": "Response Optimized for Cyber Threats"
}

# Priority rules for mapping to DII 4.0
MAPPING_RULES = {
    "priority_1": {
        "description": "Direct archetype mapping based on recovery focus",
        "mappings": {
            "ARCHETYPE_01": {"model": "A-RC", "confidence": 0.95, "reason": "Pure recovery focus across all threats"},
            "ARCHETYPE_02": {"model": "C-RY", "confidence": 0.95, "reason": "Cyber-specific recovery optimization"},
            "ARCHETYPE_03": {"model": "A-RS", "confidence": 0.90, "reason": "Balanced resilience approach"},
            "ARCHETYPE_04": {"model": "C-RS", "confidence": 0.90, "reason": "Cyber resilience focus"},
            "ARCHETYPE_05": {"model": "A-RG", "confidence": 0.85, "reason": "Response-oriented approach"},
            "ARCHETYPE_06": {"model": "C-RG", "confidence": 0.85, "reason": "Cyber response specialization"}
        }
    },
    "priority_2": {
        "description": "Business model modifiers",
        "modifiers": {
            "B2B": {"confidence_boost": 0.05, "cyber_preference": 0.2},
            "B2C": {"confidence_boost": 0.03, "all_threats_preference": 0.1},
            "B2G": {"confidence_boost": 0.08, "resilience_preference": 0.15}
        }
    },
    "priority_3": {
        "description": "Cloud adoption level impact",
        "modifiers": {
            "HIGH": {"cyber_boost": 0.1, "recovery_boost": 0.05},
            "MEDIUM": {"resilience_boost": 0.05},
            "LOW": {"response_boost": 0.05, "all_threats_boost": 0.05}
        }
    },
    "priority_4": {
        "description": "Sector-specific adjustments",
        "sector_preferences": {
            "FINANCIAL": {"models": ["C-RS", "C-RY"], "boost": 0.1},
            "HEALTHCARE": {"models": ["A-RS", "A-RC"], "boost": 0.1},
            "RETAIL": {"models": ["C-RG", "A-RG"], "boost": 0.08},
            "TECHNOLOGY": {"models": ["C-RY", "C-RS"], "boost": 0.12},
            "MANUFACTURING": {"models": ["A-RC", "A-RS"], "boost": 0.08},
            "GOVERNMENT": {"models": ["A-RS", "C-RS"], "boost": 0.15}
        }
    }
}

def read_excel_with_openpyxl(file_path):
    """Try to read Excel file using openpyxl."""
    try:
        from openpyxl import load_workbook
        
        wb = load_workbook(filename=file_path, read_only=True)
        data = {}
        
        for sheet_name in wb.sheetnames:
            print(f"Loading sheet: {sheet_name}")
            ws = wb[sheet_name]
            
            # Get headers from first row
            headers = []
            for cell in ws[1]:
                headers.append(str(cell.value) if cell.value else '')
            
            # Read data rows
            rows = []
            for row_idx, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
                if any(cell is not None for cell in row):  # Skip empty rows
                    row_dict = {}
                    for col_idx, value in enumerate(row):
                        if col_idx < len(headers) and headers[col_idx]:
                            row_dict[headers[col_idx]] = value
                    rows.append(row_dict)
            
            data[sheet_name] = rows
        
        wb.close()
        return data
        
    except ImportError:
        print("openpyxl not available")
        return None
    except Exception as e:
        print(f"Error reading Excel file with openpyxl: {e}")
        return None

def find_dim_clients_sheet(data):
    """Find the Dim_Clients sheet or similar."""
    for sheet_name in data.keys():
        if 'dim_clients' in sheet_name.lower() or 'clients' in sheet_name.lower():
            return sheet_name
    return None

def extract_unique_combinations(rows):
    """Extract unique combinations from client data."""
    required_columns = ['SECTOR', 'BUSINESS_MODEL', 'ARCHETYPE_ID', 'CLOUD_ADOPTION_LEVEL']
    
    # Group combinations
    combinations = defaultdict(int)
    
    for row in rows:
        # Try to find required columns (case-insensitive)
        combo_key = []
        for req_col in required_columns:
            value = None
            for col in row.keys():
                if req_col.lower() in col.lower():
                    value = row[col]
                    break
            combo_key.append(str(value) if value else 'UNKNOWN')
        
        combo_tuple = tuple(combo_key)
        combinations[combo_tuple] += 1
    
    # Convert to list of dictionaries
    results = []
    for combo, count in combinations.items():
        results.append({
            'SECTOR': combo[0],
            'BUSINESS_MODEL': combo[1],
            'ARCHETYPE_ID': combo[2],
            'CLOUD_ADOPTION_LEVEL': combo[3],
            'CLIENT_COUNT': count
        })
    
    # Sort by count
    results.sort(key=lambda x: x['CLIENT_COUNT'], reverse=True)
    
    return results

def calculate_dii_mapping(row):
    """Calculate DII 4.0 mapping for a specific combination."""
    mapping_scores = {model: 0.0 for model in DII_4_MODELS.keys()}
    confidence_factors = []
    reasoning = []
    
    # Priority 1: Archetype-based mapping
    archetype = row.get('ARCHETYPE_ID', 'UNKNOWN')
    if archetype in MAPPING_RULES['priority_1']['mappings']:
        mapping_info = MAPPING_RULES['priority_1']['mappings'][archetype]
        primary_model = mapping_info['model']
        mapping_scores[primary_model] = mapping_info['confidence']
        confidence_factors.append(mapping_info['confidence'])
        reasoning.append(f"Archetype {archetype}: {mapping_info['reason']}")
    else:
        # Default mapping if archetype unknown
        mapping_scores['A-RS'] = 0.5  # Default to balanced resilience
        confidence_factors.append(0.5)
        reasoning.append("Unknown archetype - defaulting to balanced approach")
    
    # Priority 2: Business model modifier
    business_model = row.get('BUSINESS_MODEL', 'UNKNOWN')
    if business_model in MAPPING_RULES['priority_2']['modifiers']:
        modifier = MAPPING_RULES['priority_2']['modifiers'][business_model]
        
        # Apply cyber preference
        if 'cyber_preference' in modifier:
            for model in ['C-RY', 'C-RS', 'C-RG']:
                mapping_scores[model] *= (1 + modifier['cyber_preference'])
        
        # Apply all threats preference
        if 'all_threats_preference' in modifier:
            for model in ['A-RC', 'A-RS', 'A-RG']:
                mapping_scores[model] *= (1 + modifier['all_threats_preference'])
        
        # Apply resilience preference
        if 'resilience_preference' in modifier:
            for model in ['A-RS', 'C-RS']:
                mapping_scores[model] *= (1 + modifier['resilience_preference'])
        
        confidence_factors.append(modifier['confidence_boost'])
        reasoning.append(f"Business model {business_model} adjustments applied")
    
    # Priority 3: Cloud adoption impact
    cloud_level = row.get('CLOUD_ADOPTION_LEVEL', 'UNKNOWN')
    if cloud_level in MAPPING_RULES['priority_3']['modifiers']:
        modifier = MAPPING_RULES['priority_3']['modifiers'][cloud_level]
        
        if 'cyber_boost' in modifier:
            for model in ['C-RY', 'C-RS', 'C-RG']:
                mapping_scores[model] *= (1 + modifier['cyber_boost'])
        
        if 'recovery_boost' in modifier:
            for model in ['A-RC', 'C-RY']:
                mapping_scores[model] *= (1 + modifier['recovery_boost'])
        
        if 'resilience_boost' in modifier:
            for model in ['A-RS', 'C-RS']:
                mapping_scores[model] *= (1 + modifier['resilience_boost'])
        
        if 'response_boost' in modifier:
            for model in ['A-RG', 'C-RG']:
                mapping_scores[model] *= (1 + modifier['response_boost'])
        
        if 'all_threats_boost' in modifier:
            for model in ['A-RC', 'A-RS', 'A-RG']:
                mapping_scores[model] *= (1 + modifier['all_threats_boost'])
        
        reasoning.append(f"Cloud adoption level {cloud_level} impacts applied")
    
    # Priority 4: Sector-specific adjustments
    sector = row.get('SECTOR', 'UNKNOWN')
    if sector in MAPPING_RULES['priority_4']['sector_preferences']:
        pref = MAPPING_RULES['priority_4']['sector_preferences'][sector]
        for model in pref['models']:
            mapping_scores[model] *= (1 + pref['boost'])
        reasoning.append(f"Sector {sector} preferences applied")
    
    # Normalize scores
    total_score = sum(mapping_scores.values())
    if total_score > 0:
        for model in mapping_scores:
            mapping_scores[model] /= total_score
    
    # Determine primary recommendation
    primary_model = max(mapping_scores, key=mapping_scores.get)
    
    # Calculate overall confidence
    base_confidence = statistics.mean(confidence_factors) if confidence_factors else 0.5
    score_confidence = mapping_scores[primary_model]
    overall_confidence = (base_confidence + score_confidence) / 2
    
    return {
        'primary_model': primary_model,
        'model_description': DII_4_MODELS[primary_model],
        'confidence': round(overall_confidence, 3),
        'all_scores': {k: round(v, 3) for k, v in mapping_scores.items()},
        'reasoning': ' | '.join(reasoning)
    }

def extract_recovery_agility(data):
    """Extract Recovery Agility indicators from other sheets."""
    recovery_indicators = {}
    
    # Look for sheets that might contain recovery agility data
    potential_sheets = ['Recovery', 'Agility', 'Metrics', 'KPIs', 'Performance']
    
    for sheet_name in data.keys():
        if any(keyword.lower() in sheet_name.lower() for keyword in potential_sheets):
            rows = data[sheet_name]
            if rows:
                print(f"\nAnalyzing sheet '{sheet_name}' for recovery indicators:")
                
                # Get all unique columns from rows
                all_columns = set()
                for row in rows[:10]:  # Sample first 10 rows
                    all_columns.update(row.keys())
                
                print(f"Columns: {list(all_columns)[:10]}...")  # Show first 10 columns
                
                # Look for recovery-related columns
                recovery_cols = [col for col in all_columns if any(
                    keyword in str(col).upper() for keyword in ['RECOVERY', 'AGILITY', 'RTO', 'RPO', 'MTTR']
                )]
                
                if recovery_cols:
                    sample_data = []
                    for row in rows[:5]:  # Get first 5 rows as sample
                        sample_row = {col: row.get(col, '') for col in recovery_cols}
                        sample_data.append(sample_row)
                    
                    recovery_indicators[sheet_name] = {
                        'columns': recovery_cols,
                        'sample_data': sample_data
                    }
    
    return recovery_indicators

def create_mapping_matrix(grouped_data):
    """Create comprehensive mapping matrix."""
    results = []
    
    for row in grouped_data:
        mapping = calculate_dii_mapping(row)
        
        result = {
            'combination': {
                'sector': row.get('SECTOR', 'N/A'),
                'business_model': row.get('BUSINESS_MODEL', 'N/A'),
                'archetype_id': row.get('ARCHETYPE_ID', 'N/A'),
                'cloud_adoption_level': row.get('CLOUD_ADOPTION_LEVEL', 'N/A')
            },
            'client_count': row.get('CLIENT_COUNT', 0),
            'dii_4_mapping': mapping
        }
        
        results.append(result)
    
    return results

def calculate_summary_stats(mapping_matrix):
    """Calculate summary statistics."""
    model_counts = Counter()
    confidence_levels = []
    sector_distribution = Counter()
    
    for item in mapping_matrix:
        model_counts[item['dii_4_mapping']['primary_model']] += item['client_count']
        confidence_levels.append(item['dii_4_mapping']['confidence'])
        sector_distribution[item['combination']['sector']] += item['client_count']
    
    return {
        'model_distribution': dict(model_counts),
        'average_confidence': round(statistics.mean(confidence_levels), 3) if confidence_levels else 0,
        'confidence_std_dev': round(statistics.stdev(confidence_levels), 3) if len(confidence_levels) > 1 else 0,
        'top_sectors': dict(sector_distribution.most_common(5))
    }

def generate_report(mapping_matrix, recovery_indicators, output_file):
    """Generate comprehensive report."""
    report = {
        'metadata': {
            'generated_date': datetime.now().isoformat(),
            'total_combinations': len(mapping_matrix),
            'total_clients': sum(item['client_count'] for item in mapping_matrix)
        },
        'mapping_matrix': mapping_matrix,
        'recovery_indicators': recovery_indicators,
        'summary_statistics': calculate_summary_stats(mapping_matrix)
    }
    
    # Save as JSON
    with open(output_file + '.json', 'w') as f:
        json.dump(report, f, indent=2)
    
    # Generate readable text report
    with open(output_file + '.txt', 'w') as f:
        f.write("DII 4.0 MAPPING ANALYSIS REPORT\n")
        f.write("=" * 80 + "\n\n")
        f.write(f"Generated: {report['metadata']['generated_date']}\n")
        f.write(f"Total Unique Combinations: {report['metadata']['total_combinations']}\n")
        f.write(f"Total Clients Analyzed: {report['metadata']['total_clients']}\n\n")
        
        f.write("MAPPING MATRIX\n")
        f.write("-" * 80 + "\n\n")
        
        for i, item in enumerate(mapping_matrix, 1):
            f.write(f"Combination #{i}\n")
            f.write(f"Sector: {item['combination']['sector']}\n")
            f.write(f"Business Model: {item['combination']['business_model']}\n")
            f.write(f"Archetype: {item['combination']['archetype_id']}\n")
            f.write(f"Cloud Adoption: {item['combination']['cloud_adoption_level']}\n")
            f.write(f"Client Count: {item['client_count']}\n")
            f.write(f"Recommended DII 4.0 Model: {item['dii_4_mapping']['primary_model']} - {item['dii_4_mapping']['model_description']}\n")
            f.write(f"Confidence: {item['dii_4_mapping']['confidence']:.1%}\n")
            f.write(f"Reasoning: {item['dii_4_mapping']['reasoning']}\n")
            f.write(f"All Model Scores: {item['dii_4_mapping']['all_scores']}\n")
            f.write("-" * 40 + "\n\n")
        
        f.write("\nSUMMARY STATISTICS\n")
        f.write("-" * 80 + "\n")
        for key, value in report['summary_statistics'].items():
            f.write(f"{key}: {value}\n")
        
        if recovery_indicators:
            f.write("\n\nRECOVERY AGILITY INDICATORS\n")
            f.write("-" * 80 + "\n")
            for sheet_name, indicators in recovery_indicators.items():
                f.write(f"\nSheet: {sheet_name}\n")
                f.write(f"Recovery-related columns: {indicators['columns']}\n")
                f.write("Sample data:\n")
                for row in indicators['sample_data']:
                    f.write(f"  {row}\n")

def main():
    """Main execution function."""
    file_path = './data/MasterDatabaseV3.1.xlsx'
    
    print(f"Loading data from {file_path}...")
    
    # Try to read Excel file
    data = read_excel_with_openpyxl(file_path)
    
    if data is None:
        print("Failed to load Excel file. Please ensure openpyxl is installed or convert to CSV.")
        return
    
    print(f"\nFound {len(data)} sheets in the workbook")
    print(f"Sheet names: {list(data.keys())}")
    
    # Find Dim_Clients sheet
    dim_clients_sheet = find_dim_clients_sheet(data)
    
    if dim_clients_sheet is None:
        print("Error: Could not find Dim_Clients sheet")
        print("Available sheets:", list(data.keys()))
        return
    
    print(f"\nUsing sheet: {dim_clients_sheet}")
    dim_clients_data = data[dim_clients_sheet]
    print(f"Number of rows: {len(dim_clients_data)}")
    
    if dim_clients_data:
        # Show sample columns from first row
        print(f"Sample columns: {list(dim_clients_data[0].keys())[:10]}")
    
    # Extract unique combinations
    print("\nExtracting unique combinations...")
    grouped_data = extract_unique_combinations(dim_clients_data)
    print(f"Found {len(grouped_data)} unique combinations")
    
    # Extract recovery agility indicators
    print("\nExtracting recovery agility indicators...")
    recovery_indicators = extract_recovery_agility(data)
    
    # Create mapping matrix
    print("\nCreating DII 4.0 mapping matrix...")
    mapping_matrix = create_mapping_matrix(grouped_data)
    
    # Generate report
    output_file = 'dii_4_mapping_analysis'
    print(f"\nGenerating report: {output_file}...")
    generate_report(mapping_matrix, recovery_indicators, output_file)
    
    print("\nAnalysis complete!")
    print(f"Results saved to:")
    print(f"  - {output_file}.json (detailed data)")
    print(f"  - {output_file}.txt (readable report)")

if __name__ == "__main__":
    main()